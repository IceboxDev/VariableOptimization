"""Shared fixtures: a small sanitized xlsx clone exercising every edge case
the live sheet has produced — mid-column blanks, the -1 score sentinel, an
Overlap row with no weight, N/A players, and overnight game durations."""

import datetime
from pathlib import Path

import openpyxl
import pytest

from variableoptimization import constants
from variableoptimization.snapshot import Snapshot
from variableoptimization.sources import XlsxSource

GAMES = [
    # (date, duration, score, anomaly, players)
    (datetime.date(2024, 1, 5), "18:30 - 20:15", 40, False, ["Alice", "Bob", "Carol"]),
    (datetime.date(2024, 1, 12), None, None, False, ["Alice", "Dave"]),  # blank score mid-column
    (datetime.date(2024, 1, 19), "23:30 - 00:45", -1, False, ["Bob", "Carol", "Dave"]),
    (datetime.date(2024, 1, 26), None, 55, True, ["Alice", "Bob", "N/A"]),
    (datetime.date(2024, 2, 2), None, 33, False, ["Carol", "Dave"]),
    (datetime.date(2024, 2, 9), None, None, None, ["Alice", "Carol"]),  # upcoming game
]

WEIGHTS = [
    ("Alice", 0.5),
    ("Bob", 0.25),
    ("Overlap", None),  # the live-sheet bug: Overlap row with an empty weight
]


@pytest.fixture(scope="session")
def fixture_xlsx(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("data") / "fixture.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Game Library"  # a decoy sheet, like the real file
    sheet = workbook.create_sheet(constants.WORKSHEET_NAME)

    for offset, (date, duration, score, anomaly, players) in enumerate(GAMES):
        row = constants.GAMES_FIRST_ROW + offset
        sheet.cell(row, constants.COL_DATE, date)
        if duration is not None:
            sheet.cell(row, constants.COL_DURATION, duration)
        if score is not None:
            sheet.cell(row, constants.COL_SCORE, score)
        if anomaly is not None:
            sheet.cell(row, constants.COL_ANOMALY, anomaly)
        for player_offset, player in enumerate(players):
            sheet.cell(row, constants.COL_PLAYERS_FIRST + player_offset, player)

    for offset, (name, weight) in enumerate(WEIGHTS):
        row = constants.WEIGHTS_FIRST_ROW + offset
        sheet.cell(row, constants.COL_WEIGHT_NAME, name)
        if weight is not None:
            sheet.cell(row, constants.COL_WEIGHT_VALUE, weight)

    workbook.save(path)
    return path


@pytest.fixture()
def fixture_snapshot(fixture_xlsx) -> Snapshot:
    return XlsxSource(fixture_xlsx).fetch()

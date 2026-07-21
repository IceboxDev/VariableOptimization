"""Local xlsx data source — reads an exported clone of the Google Sheet.

Produces the exact same Snapshot as :class:`SheetsSource`, with cell values
canonicalised to the sheet's string conventions (dates as MM/DD/YY, booleans
as TRUE/FALSE), so downstream code cannot tell the sources apart.
"""

import datetime
import logging
from pathlib import Path

import openpyxl

from .. import constants
from ..snapshot import GameRecord, Snapshot, WeightRecord

log = logging.getLogger(__name__)


def _canonical(value: object) -> str:
    """Convert an openpyxl cell value to the sheet's string convention."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime(constants.DATE_FORMAT)
    if isinstance(value, datetime.time):
        return value.strftime(constants.TIME_FORMAT)
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


class XlsxSource:
    def __init__(
        self,
        path: str | Path,
        worksheet_name: str = constants.WORKSHEET_NAME,
    ) -> None:
        self._path = Path(path)
        self._worksheet_name = worksheet_name
        if not self._path.is_file():
            raise FileNotFoundError(f"xlsx file not found: {self._path}")

    def fetch(self) -> Snapshot:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        try:
            if self._worksheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet {self._worksheet_name!r} not found in {self._path}; "
                    f"available: {workbook.sheetnames}"
                )
            sheet = workbook[self._worksheet_name]

            games = []
            for row in range(constants.GAMES_FIRST_ROW, sheet.max_row + 1):
                date = _canonical(sheet.cell(row, constants.COL_DATE).value)
                players = tuple(
                    name
                    for column in range(
                        constants.COL_PLAYERS_FIRST, constants.COL_PLAYERS_LAST + 1
                    )
                    if (name := _canonical(sheet.cell(row, column).value))
                )
                if not date and not players:
                    continue
                games.append(
                    GameRecord(
                        row=row,
                        date=date,
                        duration=_canonical(
                            sheet.cell(row, constants.COL_DURATION).value
                        ),
                        score=_canonical(sheet.cell(row, constants.COL_SCORE).value),
                        anomaly=_canonical(
                            sheet.cell(row, constants.COL_ANOMALY).value
                        ),
                        players=players,
                    )
                )

            weights = []
            for row in range(constants.WEIGHTS_FIRST_ROW, sheet.max_row + 1):
                name = _canonical(sheet.cell(row, constants.COL_WEIGHT_NAME).value)
                if name:
                    weights.append(
                        WeightRecord(
                            name=name,
                            weight=_canonical(
                                sheet.cell(row, constants.COL_WEIGHT_VALUE).value
                            ),
                        )
                    )
        finally:
            workbook.close()

        log.debug(
            "Read %d game rows, %d weight rows from %s",
            len(games), len(weights), self._path,
        )
        return Snapshot(games=tuple(games), weights=tuple(weights))
